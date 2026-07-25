"""Shared email pipeline infrastructure.

Reusable components for email analysis pipelines:
- LLM call with retry and model fallback
- Image signature detection and caching
- Attachment content extraction (PDF, image, spreadsheet)

Pipeline-specific logic (prompts, interpretation, DB writes) stays in
each pipeline's own module (e.g. supplier_quote_pipeline.py).
"""

import base64
import hashlib
import logging
import os
import time

from config.settings import Config

logger = logging.getLogger(__name__)

# Default fallback model — fast, non-thinking, reliable
FALLBACK_MODEL = "gemini-2.0-flash"


# ---------------------------------------------------------------------------
# Model resolution
# ---------------------------------------------------------------------------

def get_pipeline_model(pipeline: str, step: str) -> str:
    """Resolve the LLM model for a pipeline step.

    Resolution order:
      1. Step-specific env var: {PIPELINE}_{STEP}_MODEL (e.g. QUOTE_CLASSIFY_MODEL)
      2. Pipeline-level env var: {PIPELINE}_PIPELINE_MODEL (e.g. QUOTE_PIPELINE_MODEL)
      3. Config.DEFAULT_MODEL

    Args:
        pipeline: Pipeline name prefix, e.g. "QUOTE", "CUSTOMER_REQUEST"
        step: Step name, e.g. "classify", "extract", "interpret"
    """
    # Step-specific: QUOTE_CLASSIFY_MODEL
    step_var = f"{pipeline}_{step.upper()}_MODEL"
    step_model = os.getenv(step_var, "")
    if step_model:
        return step_model

    # Pipeline-level: QUOTE_PIPELINE_MODEL
    pipeline_var = f"{pipeline}_PIPELINE_MODEL"
    pipeline_model = os.getenv(pipeline_var, "")
    if pipeline_model:
        return pipeline_model

    return Config.DEFAULT_MODEL


# ---------------------------------------------------------------------------
# LLM call with retry + model fallback
# ---------------------------------------------------------------------------

def llm_call_with_retry(
    pipeline: str,
    step: str,
    contents,
    temperature: float = 0.1,
    timeout: int = 120000,
):
    """Call Gemini with retry on transient errors (504, 503, DEADLINE_EXCEEDED).

    Retry strategy:
      1. Try primary model for the pipeline/step
      2. Retry primary model once (transient 504s)
      3. Fall back to FALLBACK_MODEL

    Args:
        pipeline: Pipeline name prefix (e.g. "QUOTE")
        step: Step name (e.g. "classify", "extract", "interpret")
        contents: Gemini contents (string, list of Parts, etc.)
        temperature: LLM temperature
        timeout: HTTP timeout in milliseconds

    Returns the response object. Raises on permanent failure after all retries.
    """
    from google import genai as _genai
    from google.genai import types as _types

    primary_model = get_pipeline_model(pipeline, step)
    models_to_try = [primary_model, primary_model, FALLBACK_MODEL]

    last_error = None
    for attempt, model in enumerate(models_to_try):
        try:
            client = _genai.Client(http_options={"timeout": timeout})
            response = client.models.generate_content(
                model=model,
                contents=contents,
                config=_types.GenerateContentConfig(temperature=temperature),
            )
            if attempt > 0:
                logger.info(
                    f"[email-pipeline] {pipeline}/{step}: succeeded on "
                    f"attempt {attempt + 1} (model={model})"
                )
            return response
        except Exception as e:
            error_str = str(e)
            is_transient = any(
                code in error_str
                for code in (
                    "504", "503", "DEADLINE_EXCEEDED",
                    "UNAVAILABLE", "RESOURCE_EXHAUSTED",
                )
            )
            if not is_transient:
                raise  # permanent error — don't retry
            last_error = e
            logger.warning(
                f"[email-pipeline] {pipeline}/{step}: attempt {attempt + 1} "
                f"failed (model={model}): {e}"
            )
            if attempt < len(models_to_try) - 1:
                time.sleep(2 ** attempt)  # 1s, 2s backoff

    raise last_error  # all retries exhausted


# ---------------------------------------------------------------------------
# Database session helper
# ---------------------------------------------------------------------------

def _get_session():
    """Get a synchronous SQLAlchemy session."""
    from includes.dashboard.database import get_session
    return get_session()


# ---------------------------------------------------------------------------
# Image signature detection & caching
# ---------------------------------------------------------------------------

def check_image_signature(image_bytes: bytes) -> str | None:
    """Check if image bytes match a known signature.

    Returns 'signature' or 'quote_content' if known, None if unknown.
    """
    sha = hashlib.sha256(image_bytes).hexdigest()
    session = _get_session()
    try:
        from includes.dashboard.models import KnownImageSignature
        record = session.query(KnownImageSignature).filter(
            KnownImageSignature.sha256 == sha
        ).first()
        return record.classification if record else None
    finally:
        session.close()


def store_image_signature(
    image_bytes: bytes,
    classification: str,
    filename: str | None = None,
    source_email_id: int | None = None,
) -> None:
    """Store image hash and classification for future lookups."""
    sha = hashlib.sha256(image_bytes).hexdigest()
    session = _get_session()
    try:
        from includes.dashboard.models import KnownImageSignature
        existing = session.query(KnownImageSignature).filter(
            KnownImageSignature.sha256 == sha
        ).first()
        if existing:
            return  # already known
        from datetime import datetime, timezone
        record = KnownImageSignature(
            sha256=sha,
            classification=classification,
            sample_filename=filename,
            source_email_id=source_email_id,
            size_bytes=len(image_bytes),
            created_at=datetime.now(timezone.utc),
        )
        session.add(record)
        session.commit()
    except Exception as e:
        session.rollback()
        logger.warning(f"Failed to store image signature: {e}")
    finally:
        session.close()


def classify_image_content(
    image_bytes: bytes,
    mime_type: str,
    pipeline: str = "QUOTE",
) -> str:
    """Ask the LLM whether an image is a signature/logo or meaningful content.

    Returns 'signature' or 'quote_content'.
    Uses the extract model for the given pipeline.
    """
    from google.genai import types as _types

    try:
        response = llm_call_with_retry(
            pipeline=pipeline,
            step="extract",
            contents=[
                _types.Part.from_bytes(data=image_bytes, mime_type=mime_type),
                (
                    "Is this image a corporate logo, email signature, banner, "
                    "social media icon, or decorative element? Or does it contain "
                    "meaningful content like a quotation, price list, invoice, "
                    "product specifications, or tabular data?\n\n"
                    "Reply with ONLY one word: 'signature' or 'content'"
                ),
            ],
            temperature=0.0,
            timeout=15000,
        )
        answer = (response.text or "").strip().lower()
        if "signature" in answer:
            return "signature"
        return "quote_content"
    except Exception as e:
        logger.warning(f"Image classification failed, assuming quote_content: {e}")
        return "quote_content"  # err on side of inclusion


def triage_image(
    image_bytes: bytes,
    mime_type: str,
    filename: str | None = None,
    source_email_id: int | None = None,
    pipeline: str = "QUOTE",
) -> str:
    """Full image triage: check cache → classify if unknown → cache result.

    Returns 'signature' or 'quote_content'.
    """
    # 1. Check cache
    cached = check_image_signature(image_bytes)
    if cached is not None:
        if cached == "signature":
            logger.debug(f"[email-pipeline] Skipping known signature: {filename}")
        return cached

    # 2. Classify unknown image
    classification = classify_image_content(image_bytes, mime_type, pipeline)

    # 3. Cache for future
    store_image_signature(image_bytes, classification, filename, source_email_id)
    if classification == "signature":
        logger.debug(f"[email-pipeline] New signature detected & cached: {filename}")
    return classification


# ---------------------------------------------------------------------------
# Gmail attachment fetch
# ---------------------------------------------------------------------------

def fetch_gmail_attachment_bytes(
    user_email: str, message_id: str, attachment_id: str
) -> bytes | None:
    """Fetch raw attachment bytes from Gmail API."""
    try:
        from includes.gmail import get_gmail_client
        service = get_gmail_client(user_email)
        attachment = (
            service.users().messages().attachments()
            .get(userId="me", messageId=message_id, id=attachment_id)
            .execute()
        )
        return base64.urlsafe_b64decode(attachment["data"])
    except Exception as e:
        logger.warning(f"Failed to fetch attachment {message_id}/{attachment_id}: {e}")
        return None


# ---------------------------------------------------------------------------
# Attachment content extraction
# ---------------------------------------------------------------------------

def extract_pdf_content(
    pdf_bytes: bytes,
    filename: str,
    pipeline: str = "QUOTE",
) -> str:
    """Extract text and tabular data from a PDF via Gemini."""
    from google.genai import types as _types

    try:
        response = llm_call_with_retry(
            pipeline=pipeline,
            step="extract",
            contents=[
                _types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
                (
                    "Extract ALL text, pricing information, part numbers, quantities, "
                    "and tabular data from this document. Present the data as Markdown "
                    "tables. Include:\n"
                    "- Item descriptions and part numbers\n"
                    "- Unit prices and totals with currency\n"
                    "- Quantities and units of measure\n"
                    "- Shipping costs, lead times, payment terms if mentioned\n"
                    "- Any notes or conditions\n\n"
                    "If the document contains multiple tables, reproduce each one. "
                    "Preserve the original structure as faithfully as possible."
                ),
            ],
            temperature=0.1,
            timeout=120000,
        )
        if not response.text:
            candidates = getattr(response, "candidates", None)
            if candidates and candidates[0].finish_reason:
                logger.warning(
                    f"Gemini PDF extraction empty for {filename}: "
                    f"finish_reason={candidates[0].finish_reason}"
                )
            return "*[No content extracted from PDF]*"
        return response.text
    except Exception as e:
        logger.error(f"Gemini PDF extraction failed for {filename}: {e}")
        return f"*[PDF extraction failed: {e}]*"


def extract_image_content(
    image_bytes: bytes,
    filename: str,
    mime_type: str,
    pipeline: str = "QUOTE",
) -> str:
    """Extract text and tabular data from an image via Gemini OCR."""
    from google.genai import types as _types

    try:
        response = llm_call_with_retry(
            pipeline=pipeline,
            step="extract",
            contents=[
                _types.Part.from_bytes(data=image_bytes, mime_type=mime_type),
                (
                    "Extract all text, pricing, and tabular data from this image. "
                    "If it's a quotation or price list, present as a Markdown table. "
                    "If it's a product spec sheet, extract key specifications. "
                    "If it's general text, reproduce it faithfully."
                ),
            ],
            temperature=0.1,
            timeout=60000,
        )
        return response.text or "*[No content extracted]*"
    except Exception as e:
        logger.error(f"Gemini image extraction failed for {filename}: {e}")
        return f"*[Image extraction failed: {e}]*"


def extract_spreadsheet_content(data: bytes, filename: str, mime_type: str) -> str:
    """Extract spreadsheet content. CSV parsed directly; Excel via openpyxl.

    No LLM call — purely local parsing.
    """
    if filename.lower().endswith(".csv"):
        try:
            text = data.decode("utf-8", errors="replace")
            return f"```csv\n{text[:5000]}\n```"
        except Exception:
            pass

    try:
        import io
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            parts.append(f"## Sheet: {sheet_name}\n")
            header = rows[0]
            col_names = [str(c) if c is not None else "" for c in header]
            parts.append("| " + " | ".join(col_names) + " |")
            parts.append("| " + " | ".join(["---"] * len(col_names)) + " |")
            for row in rows[1:200]:
                cells = [str(c) if c is not None else "" for c in row]
                parts.append("| " + " | ".join(cells) + " |")
        wb.close()
        result_text = "\n".join(parts)
        if not result_text.strip():
            return "*[Spreadsheet appears empty]*"
        return result_text[:8000]
    except Exception as e:
        logger.error(f"Local spreadsheet extraction failed for {filename}: {e}")
        return f"*[Spreadsheet extraction failed: {e}]*"
