"""Supplier quote pipeline — classify, extract, and interpret supplier quote emails.

Three-stage pipeline:
  1. classify_supplier_email — triage: is this a quote response?
  2. extract_email_content — gather: fetch body + process attachments via Gemini
  3. interpret_quote_response — apply: match extracted data to RFQ items and update

Tools are created via create_supplier_quote_tools(user_id) and registered with the agent.
"""

import asyncio
import base64
import json
import logging
import re
from typing import Optional

from langchain_core.tools import tool

from config.settings import Config

# Model for pipeline LLM calls — falls back to DEFAULT_MODEL if not set
_PIPELINE_MODEL = Config.QUOTE_PIPELINE_MODEL or Config.DEFAULT_MODEL

logger = logging.getLogger(__name__)

# Quote indicator keywords for classification
_QUOTE_KEYWORDS = re.compile(
    r'(quot(e|ation)|pricing|price list|unit price|lead\s*time|ex[\s-]?works|'
    r'fob|cif|cfr|validity|payment\s*terms|pro[\s-]?forma|offer)',
    re.IGNORECASE,
)

# Currency patterns
_CURRENCY_RE = re.compile(
    r'(\$|€|£|¥|USD|AUD|EUR|GBP|JPY|SGD|NZD)\s*[\d,]+\.?\d*',
    re.IGNORECASE,
)


def _get_session():
    """Get a synchronous SQLAlchemy session."""
    from includes.dashboard.database import get_session
    return get_session()


def _get_email_tracking(session, email_tracking_id: int):
    """Fetch an EmailTracking record by ID."""
    from includes.dashboard.models import EmailTracking
    return session.query(EmailTracking).filter(EmailTracking.id == email_tracking_id).first()


def _fetch_gmail_attachment_bytes(user_email: str, message_id: str, attachment_id: str) -> bytes | None:
    """Fetch raw attachment bytes from Gmail API."""
    try:
        from includes.gmail import get_gmail_client
        service = get_gmail_client(user_email)
        attachment = (
            service.users().messages().attachments()
            .get(userId="me", messageId=message_id, id=attachment_id)
            .execute()
        )
        return base64.urlsafe_b64decode(attachment["data"])
    except Exception as e:
        logger.warning(f"Failed to fetch attachment {message_id}/{attachment_id}: {e}")
        return None


def _now_iso() -> str:
    """Return current UTC time as ISO string."""
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()


def _save_pipeline_result(email_tracking_id: int, result: dict) -> None:
    """Persist pipeline result to the email_tracking record."""
    session = _get_session()
    try:
        tracking = _get_email_tracking(session, email_tracking_id)
        if not tracking:
            print(f"[quote-pipeline] #{email_tracking_id}: cannot save result — email not found", flush=True)
            logger.warning(f"[quote-pipeline] #{email_tracking_id}: cannot save result — email not found")
            return
        tracking.supplier_pipeline_result = result
        session.commit()
        print(f"[quote-pipeline] #{email_tracking_id}: result saved ({result.get('classification', '?')})", flush=True)
        logger.info(f"[quote-pipeline] #{email_tracking_id}: result saved ({result.get('classification', '?')})")
    except Exception as e:
        session.rollback()
        print(f"[quote-pipeline] #{email_tracking_id}: failed to save result — {e}", flush=True)
        logger.warning(f"[quote-pipeline] #{email_tracking_id}: failed to save result — {e}")
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Stage 1: Classify (LLM-based)
# ---------------------------------------------------------------------------

_SUPPLIER_CLASSIFY_PROMPT = """\
You are an email classifier for a procurement team. Your job is to determine whether
an inbound email contains a supplier quote/pricing response for an RFQ (Request for Quote).

Classify the email into ONE of these categories:

- **quote_response**: The email contains pricing, a quote, a proforma invoice, or references
  an attached quote document. This includes partial quotes, revised quotes, or price indications.
  Err on the side of classifying as quote_response if there's ANY pricing or quote content.

- **clarification_required**: The supplier is asking a question that must be answered before
  they can provide a quote. Examples: "can you confirm the specs?", "what quantity do you need?",
  "do you need freight included?", "please provide dimensions/weights".

- **declined**: The supplier has explicitly declined to quote, cannot supply the product, or
  states they don't carry the item. Examples: "we are unable to quote", "product discontinued",
  "we don't stock this item", "not available in Australia".

- **acknowledgement**: The supplier has acknowledged receipt of the RFQ but provided no pricing
  or substantive response yet. Examples: "received, will revert shortly", "working on it",
  "thank you for your enquiry, we'll get back to you".

- **not_quote**: Any other email that doesn't fit the above. Examples:
  - Auto-replies / out-of-office / delivery receipts
  - General conversation / follow-ups with no new data
  - Marketing / newsletters / unrelated correspondence
  - Purchase order confirmations / shipping notifications

- **needs_review**: Genuinely ambiguous — you cannot tell from the available content.
  Use this sparingly. If in doubt between quote_response and not_quote, prefer quote_response.

IMPORTANT: If the email says they are attaching a quote/price list (even if you can't see the
attachment content), classify as quote_response.

If the email IS a quote_response, also identify which attachments (by filename) are likely
to contain the actual quote/pricing data. Look at filenames (PDFs, spreadsheets, images
named like "quote", "quotation", "price list") and body context ("see attached",
"please find attached quote"). Err on the side of inclusion — it's better to extract
a non-quote attachment than to miss quote data. Include inline images if the body
references them as containing quote data.

Respond with ONLY a JSON object:
{"classification": "quote_response"|"clarification_required"|"declined"|"acknowledgement"|"not_quote"|"needs_review", "reason": "<brief explanation>", "quote_attachments": ["filename1.pdf", "filename2.png"]}

The "quote_attachments" field is only required for quote_response. For other
classifications, omit it or use an empty list.
"""


def _classify_supplier_email_sync(email_tracking_id: int) -> dict:
    """LLM-based classification of a supplier email.

    Returns dict with:
      - classification: 'quote_response' | 'not_quote' | 'needs_review'
      - reason: explanation
      - rfq_id: linked RFQ (if any)
      - supplier_name: matched supplier (if any)
    """
    session = _get_session()
    try:
        tracking = _get_email_tracking(session, email_tracking_id)
        if not tracking:
            return {"classification": "not_quote", "reason": "Email not found"}

        result = {
            "classification": "not_quote",
            "reason": "",
            "rfq_id": tracking.rfq_token or tracking.rfq_id,
            "supplier_name": None,
            "email_tracking_id": email_tracking_id,
        }

        # Must be a received email
        if tracking.direction != "received":
            result["reason"] = f"Not an inbound email (direction={tracking.direction})"
            return result

        # Must be linked to an RFQ
        rfq_id = tracking.rfq_token or tracking.rfq_id
        if not rfq_id:
            result["reason"] = "Not linked to any RFQ"
            return result

        # Check if sender matches a shortlisted supplier
        from includes.tools.rfq_crud import _get_rfq_dict_sync
        rfq = _get_rfq_dict_sync(rfq_id)
        if not rfq:
            result["reason"] = f"RFQ '{rfq_id}' not found"
            return result

        sender_email = (tracking.sender_email or "").lower()
        sender_domain = sender_email.split("@")[-1] if "@" in sender_email else ""
        matched_supplier = None

        # Check items for shortlisted suppliers whose contacts match the sender
        for item in rfq.get("items", []):
            for supplier in (item.get("suppliers") or []):
                contacts = supplier.get("contacts") or []
                for contact in contacts:
                    contact_email = (contact.get("email") or "").lower()
                    if contact_email and (
                        contact_email == sender_email or
                        contact_email.split("@")[-1] == sender_domain
                    ):
                        matched_supplier = supplier.get("name")
                        break
                if matched_supplier:
                    break
            if matched_supplier:
                break

        # Also check supplier via the FK link
        if not matched_supplier and tracking.supplier_id:
            from includes.dashboard.models import Supplier
            supplier_obj = session.query(Supplier).filter(Supplier.id == tracking.supplier_id).first()
            if supplier_obj:
                matched_supplier = supplier_obj.name

        result["supplier_name"] = matched_supplier

        # Build context for LLM classification
        subject = tracking.subject or "(no subject)"
        body = (tracking.body_markdown or tracking.body_html or "")[:1500]
        attachment_names = []
        if tracking.attachments_json:
            for att in tracking.attachments_json:
                if not att.get("inline"):
                    attachment_names.append(att.get("filename", "unnamed"))

        email_context = f"Subject: {subject}\nFrom: {sender_email}\n"
        if attachment_names:
            email_context += f"Attachments: {', '.join(attachment_names)}\n"
        email_context += f"\nBody:\n{body}"

        # Call LLM for classification
        try:
            from google import genai as _genai
            from google.genai import types as _types

            client = _genai.Client(http_options={"timeout": 30000})
            response = client.models.generate_content(
                model=_PIPELINE_MODEL,
                contents=[
                    f"{_SUPPLIER_CLASSIFY_PROMPT}\n\n---\n\n{email_context}",
                ],
                config=_types.GenerateContentConfig(
                    temperature=0.0,
                ),
            )
            if not response.text:
                # Log WHY the response was empty (safety filter, recitation, etc.)
                candidates = getattr(response, 'candidates', None)
                if candidates and candidates[0].finish_reason:
                    reason = candidates[0].finish_reason
                    print(f"[quote-pipeline] #{email_tracking_id}: LLM empty response — finish_reason={reason} (model={_PIPELINE_MODEL})", flush=True)
                    logger.warning(f"LLM classify empty for #{email_tracking_id}: finish_reason={reason} (model={_PIPELINE_MODEL})")
                else:
                    print(f"[quote-pipeline] #{email_tracking_id}: LLM empty response — no candidates", flush=True)
                    logger.warning(f"LLM classify empty for #{email_tracking_id}: no candidates returned")
                raise ValueError("LLM returned empty response")
            raw = response.text.strip()
            # Parse JSON from response (handle markdown code fences)
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError:
                print(f"[quote-pipeline] #{email_tracking_id}: LLM bad JSON — {raw[:200]}", flush=True)
                logger.warning(f"LLM classify bad JSON for #{email_tracking_id}: {raw[:200]}")
                raise
            result["classification"] = parsed.get("classification", "needs_review")
            result["reason"] = parsed.get("reason", "LLM classified")
            result["quote_attachments"] = parsed.get("quote_attachments", [])
        except Exception as e:
            # Fallback to heuristic if LLM fails
            print(f"[quote-pipeline] #{email_tracking_id}: LLM classify FAILED — {e}", flush=True)
            logger.warning(f"LLM classify failed for email {email_tracking_id}, using heuristic: {e}")
            result.update(_classify_heuristic_fallback(tracking))

        return result
    finally:
        session.close()


def _classify_heuristic_fallback(tracking) -> dict:
    """Simple heuristic fallback if LLM classification fails."""
    body = (tracking.body_markdown or "")[:2000]
    keyword_hits = len(_QUOTE_KEYWORDS.findall(body))
    currency_hits = len(_CURRENCY_RE.findall(body))

    attachment_signals = 0
    if tracking.attachments_json:
        for att in tracking.attachments_json:
            if att.get("inline"):
                continue
            fname = (att.get("filename") or "").lower()
            if any(kw in fname for kw in ("quote", "quotation", "price", "proforma", "offer")):
                attachment_signals += 1
            if fname.endswith(".pdf") or fname.endswith(".xlsx"):
                attachment_signals += 0.5

    score = keyword_hits + currency_hits + attachment_signals * 2

    if score >= 3:
        return {
            "classification": "quote_response",
            "reason": f"Heuristic fallback: {keyword_hits} keywords, {currency_hits} prices",
        }
    elif score >= 1.5:
        return {
            "classification": "needs_review",
            "reason": f"Heuristic fallback: weak signals ({score:.1f})",
        }
    else:
        return {
            "classification": "not_quote",
            "reason": "Heuristic fallback: no significant quote indicators",
        }


# ---------------------------------------------------------------------------
# Stage 2: Extract
# ---------------------------------------------------------------------------

def _extract_email_content_sync(email_tracking_id: int, quote_attachments: list[str] | None = None) -> str:
    """Synchronous extraction: fetch body + process attachments via Gemini.

    If quote_attachments is provided, only those filenames are extracted.
    Otherwise all non-inline attachments are processed.

    Returns a Markdown content bundle with body + extracted attachment content.
    """
    session = _get_session()
    try:
        tracking = _get_email_tracking(session, email_tracking_id)
        if not tracking:
            return "Error: Email not found"

        parts = []

        # Email body
        body = tracking.body_markdown or tracking.body_html or ""
        if body:
            parts.append(f"## Email Body\n\n{body}")

        # Build set of filenames to extract (case-insensitive match)
        target_filenames = None
        if quote_attachments:
            target_filenames = {f.lower().strip() for f in quote_attachments}

        # Process attachments
        if tracking.attachments_json and tracking.gmail_message_id:
            for att in tracking.attachments_json:
                filename = att.get("filename", "unknown")
                
                # If we have a filter list, only process matching attachments
                if target_filenames is not None:
                    if filename.lower().strip() not in target_filenames:
                        continue
                else:
                    # No filter — skip inline images (signatures/logos)
                    if att.get("inline"):
                        continue
                mime_type = att.get("mime_type", "")
                size = att.get("size", 0)
                att_id = att.get("gmail_attachment_id")

                if not att_id:
                    continue

                size_kb = size / 1024 if size else 0
                header = f"## Attachment: {filename} ({size_kb:.0f} KB)"

                # Fetch raw bytes from Gmail
                raw_bytes = _fetch_gmail_attachment_bytes(
                    tracking.user_email, tracking.gmail_message_id, att_id
                )
                if not raw_bytes:
                    parts.append(f"{header}\n\n*[Failed to fetch attachment]*")
                    continue

                # Process based on MIME type
                if mime_type == "application/pdf" or filename.lower().endswith(".pdf"):
                    extracted = _extract_pdf_with_gemini(raw_bytes, filename)
                    parts.append(f"{header}\n\n{extracted}")
                elif mime_type.startswith("image/"):
                    extracted = _extract_image_with_gemini(raw_bytes, filename, mime_type)
                    parts.append(f"{header}\n\n{extracted}")
                elif "spreadsheet" in mime_type or filename.lower().endswith((".xlsx", ".xls", ".csv")):
                    extracted = _extract_spreadsheet_with_gemini(raw_bytes, filename, mime_type)
                    parts.append(f"{header}\n\n{extracted}")
                else:
                    parts.append(f"{header}\n\n*[Unsupported attachment type: {mime_type}]*")

        if not parts:
            return "Error: No content found in email (no body or attachments)"

        return "\n\n---\n\n".join(parts)
    finally:
        session.close()


def _extract_pdf_with_gemini(pdf_bytes: bytes, filename: str) -> str:
    """Pass PDF bytes to Gemini for content extraction."""
    from google import genai as _genai
    from google.genai import types as _types

    try:
        client = _genai.Client(http_options={"timeout": 120000})
        response = client.models.generate_content(
            model=_PIPELINE_MODEL,
            contents=[
                _types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
                (
                    "Extract ALL pricing information, part numbers, quantities, and tabular data "
                    "from this document. Present the data as Markdown tables. Include:\n"
                    "- Item descriptions and part numbers\n"
                    "- Unit prices and totals with currency\n"
                    "- Quantities and units of measure\n"
                    "- Shipping costs, lead times, payment terms if mentioned\n"
                    "- Any notes or conditions\n\n"
                    "If the document contains multiple tables, reproduce each one. "
                    "Preserve the original structure as faithfully as possible."
                ),
            ],
            config=_types.GenerateContentConfig(
                temperature=0.1,
            ),
        )
        if not response.text:
            candidates = getattr(response, 'candidates', None)
            if candidates and candidates[0].finish_reason:
                logger.warning(f"Gemini PDF extraction empty for {filename}: finish_reason={candidates[0].finish_reason}")
            return "*[No content extracted from PDF]*"
        return response.text
    except Exception as e:
        logger.error(f"Gemini PDF extraction failed for {filename}: {e}")
        return f"*[PDF extraction failed: {e}]*"


def _extract_image_with_gemini(image_bytes: bytes, filename: str, mime_type: str) -> str:
    """Pass image to Gemini for OCR and content extraction."""
    from google import genai as _genai
    from google.genai import types as _types

    try:
        client = _genai.Client(http_options={"timeout": 60000})
        response = client.models.generate_content(
            model=_PIPELINE_MODEL,
            contents=[
                _types.Part.from_bytes(data=image_bytes, mime_type=mime_type),
                (
                    "Extract all text, pricing, and tabular data from this image. "
                    "If it's a quotation or price list, present as a Markdown table. "
                    "If it's a product spec sheet, extract key specifications. "
                    "If it's general text, reproduce it faithfully."
                ),
            ],
            config=_types.GenerateContentConfig(
                temperature=0.1,
            ),
        )
        return response.text or "*[No content extracted]*"
    except Exception as e:
        logger.error(f"Gemini image extraction failed for {filename}: {e}")
        return f"*[Image extraction failed: {e}]*"


def _extract_spreadsheet_with_gemini(data: bytes, filename: str, mime_type: str) -> str:
    """Extract spreadsheet content. For CSV, parse directly; for Excel, convert locally."""
    if filename.lower().endswith(".csv"):
        try:
            text = data.decode("utf-8", errors="replace")
            # Return first 5000 chars as-is (Markdown table conversion happens in interpret step)
            return f"```csv\n{text[:5000]}\n```"
        except Exception:
            pass

    # For Excel files, parse locally with openpyxl (Gemini doesn't accept xlsx uploads)
    try:
        import io
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            parts.append(f"## Sheet: {sheet_name}\n")
            # Format as Markdown table
            header = rows[0]
            col_names = [str(c) if c is not None else "" for c in header]
            parts.append("| " + " | ".join(col_names) + " |")
            parts.append("| " + " | ".join(["---"] * len(col_names)) + " |")
            for row in rows[1:200]:  # Cap at 200 data rows
                cells = [str(c) if c is not None else "" for c in row]
                parts.append("| " + " | ".join(cells) + " |")
        wb.close()
        result_text = "\n".join(parts)
        if not result_text.strip():
            return "*[Spreadsheet appears empty]*"
        return result_text[:8000]
    except Exception as e:
        logger.error(f"Local spreadsheet extraction failed for {filename}: {e}")
        return f"*[Spreadsheet extraction failed: {e}]*"


# ---------------------------------------------------------------------------
# Stage 3: Interpret
# ---------------------------------------------------------------------------

def _interpret_quote_sync(rfq_id: str, supplier_name: str, content_bundle: str) -> dict:
    """Use LLM to match extracted content to RFQ items and return structured quote data.

    Returns dict with:
      - quotes: [{item_line, confidence, price, currency, lead_time}, ...]
      - shipping: {cost, currency} or None
      - declined_items: [line_numbers]
      - notes: str
      - terms: str
      - warnings: [str]
    """
    from google import genai as _genai
    from google.genai import types as _types
    from includes.tools.rfq_crud import _get_rfq_dict_sync
    from includes.tools.quote_tools import _build_quotation_snapshot

    rfq = _get_rfq_dict_sync(rfq_id)
    if not rfq:
        return {"error": f"RFQ '{rfq_id}' not found"}

    # Build a concise item list for the LLM
    item_list = []
    for item in rfq.get("items", []):
        item_list.append({
            "line": item["line"],
            "description": item.get("input_description") or item.get("part_number") or "?",
            "part_number": item.get("part_number"),
            "brand": item.get("brand"),
            "quantity": item.get("quantity"),
        })

    prompt = f"""You are analyzing a supplier quote email to extract pricing data for an RFQ.

## RFQ Items (what was requested)
```json
{json.dumps(item_list, indent=2)}
```

## Supplier: {supplier_name}

## Email/Attachment Content
{content_bundle}

---

## Instructions
Match the quoted items from the email to the RFQ items above. Extract:
1. Per-item quotes: match by part number first, then by description similarity.
2. Shipping cost if mentioned.
3. Notes, terms, lead times.
4. Items explicitly declined or marked as unavailable.

For each item match, assign confidence:
- "high": exact part number match or unambiguous single-item match
- "medium": description match without exact part number
- "low": best guess based on context

Return ONLY valid JSON in this exact format:
```json
{{
  "quotes": [
    {{"item_line": 1, "confidence": "high", "price": 42.50, "currency": "AUD", "lead_time": "2 weeks"}},
  ],
  "shipping": {{"cost": 25.00, "currency": "AUD"}},
  "declined_items": [2],
  "notes": "Volume discount available",
  "terms": "Net 30",
  "warnings": ["Line 3 price is ambiguous"]
}}
```

Rules:
- `item_line` must match a line number from the RFQ items list.
- `price` is the unit price (not total).
- `currency` defaults to "AUD" if not specified in the email.
- `shipping` is null if not mentioned.
- `declined_items` lists line numbers the supplier explicitly cannot supply.
- `warnings` flags anything uncertain.
- If the email contains NO pricing data at all, return: {{"quotes": [], "shipping": null, "declined_items": [], "notes": "", "terms": "", "warnings": ["No pricing found in content"]}}
"""

    try:
        client = _genai.Client(http_options={"timeout": 120000})
        response = client.models.generate_content(
            model=_PIPELINE_MODEL,
            contents=prompt,
            config=_types.GenerateContentConfig(
                temperature=0.1,
            ),
        )
        raw_text = (response.text or "").strip()
        if not raw_text:
            candidates = getattr(response, 'candidates', None)
            if candidates and candidates[0].finish_reason:
                reason = candidates[0].finish_reason
                logger.warning(f"LLM interpret empty: finish_reason={reason}")
            else:
                logger.warning(f"LLM interpret empty: no candidates returned")
    except Exception as e:
        return {"error": f"LLM interpretation failed: {e}"}

    if not raw_text:
        return {"error": "LLM returned empty response for interpretation"}

    # Parse JSON from response
    cleaned = raw_text
    if "```json" in cleaned:
        cleaned = cleaned.split("```json", 1)[1]
    if "```" in cleaned:
        cleaned = cleaned.split("```", 1)[0]
    cleaned = cleaned.strip()

    try:
        result = json.loads(cleaned)
    except json.JSONDecodeError as e:
        return {"error": f"Failed to parse LLM response: {e}", "raw": raw_text[:500]}

    return result


def _apply_quote_data(rfq_id: str, supplier_name: str, quote_data: dict, user_id: str) -> list[str]:
    """Apply extracted quote data to the RFQ using rfq_crud functions.

    Returns list of action summaries.
    """
    from includes.tools.rfq_crud import (
        _update_supplier_sync,
        _add_supplier_sync,
        _set_supplier_meta_sync,
        _get_rfq_dict_sync,
    )

    actions = []
    rfq = _get_rfq_dict_sync(rfq_id)
    if not rfq:
        return [f"Error: RFQ '{rfq_id}' not found"]

    # Apply per-item quotes
    for quote in quote_data.get("quotes", []):
        line = quote.get("item_line")
        price = quote.get("price")
        currency = quote.get("currency", "AUD")
        lead_time = quote.get("lead_time")
        confidence = quote.get("confidence", "medium")

        if line is None or price is None:
            continue

        # Find the item
        item = next((i for i in rfq.get("items", []) if i["line"] == line), None)
        if not item:
            actions.append(f"⚠️ Line {line}: not found in RFQ")
            continue

        # Find or identify the supplier on this item
        suppliers = item.get("suppliers") or []
        supplier_entry = next(
            (s for s in suppliers if s.get("name", "").lower() == supplier_name.lower()),
            None
        )

        if supplier_entry:
            # Update existing supplier entry
            update_data = {
                "line": line,
                "name": supplier_name,
                "quote_cost": price,
                "quote_status": "quoted",
                "quote_currency": currency,
            }
            if lead_time:
                update_data["quote_leadtime"] = lead_time
            _update_supplier_sync(rfq_id, update_data, user_id)
            conf_icon = "✓" if confidence == "high" else "~" if confidence == "medium" else "?"
            actions.append(
                f"{conf_icon} Line {line}: {price} {currency}"
                f"{f' ({lead_time})' if lead_time else ''}"
            )
        else:
            # Supplier not yet on this item — add them
            add_data = {
                "line": line,
                "name": supplier_name,
                "status": "shortlisted",
                "quote_cost": price,
                "quote_status": "quoted",
                "quote_currency": currency,
            }
            if lead_time:
                add_data["quote_leadtime"] = lead_time
            _add_supplier_sync(rfq_id, add_data, user_id)
            actions.append(f"+ Line {line}: added {supplier_name} @ {price} {currency}")

    # Apply declined items
    for line in quote_data.get("declined_items", []):
        item = next((i for i in rfq.get("items", []) if i["line"] == line), None)
        if not item:
            continue
        suppliers = item.get("suppliers") or []
        supplier_entry = next(
            (s for s in suppliers if s.get("name", "").lower() == supplier_name.lower()),
            None
        )
        if supplier_entry:
            _update_supplier_sync(rfq_id, {"line": line, "name": supplier_name, "quote_status": "declined"}, user_id)
            actions.append(f"✗ Line {line}: declined")

    # Apply supplier meta (shipping, terms, notes)
    shipping = quote_data.get("shipping")
    notes = quote_data.get("notes")
    terms = quote_data.get("terms")

    if shipping or notes or terms:
        meta_data = {"name": supplier_name}
        if shipping:
            meta_data["shipping_cost"] = shipping.get("cost")
            meta_data["shipping_currency"] = shipping.get("currency", "AUD")
        if notes:
            meta_data["notes"] = notes
        if terms:
            meta_data["terms"] = terms
        _set_supplier_meta_sync(rfq_id, meta_data, user_id)
        meta_parts = []
        if shipping:
            meta_parts.append(f"shipping: {shipping.get('cost')} {shipping.get('currency', 'AUD')}")
        if terms:
            meta_parts.append(f"terms: {terms}")
        if notes:
            meta_parts.append(f"notes: {notes[:50]}")
        actions.append(f"📋 Meta: {', '.join(meta_parts)}")

    return actions


# ---------------------------------------------------------------------------
# Automated trigger — called from Gmail sync and manual email linking
# ---------------------------------------------------------------------------

def trigger_supplier_quote_pipeline(email_tracking_id: int, user_id: str = "system") -> None:
    """Run the supplier quote pipeline for an email if it's linked to both RFQ + supplier.

    Called from:
      - scripts/sync_gmail_mailboxes.py (after new received emails are committed)
      - includes/dashboard/routes/admin.py api_link_email (after manual linking)

    Runs in a background thread to avoid blocking the caller.
    Failures are logged but don't propagate.
    """
    import threading

    def _run():
        try:
            print(f"[quote-pipeline] #{email_tracking_id}: thread started", flush=True)
            logger.info(f"[quote-pipeline] #{email_tracking_id}: thread started")
            session = _get_session()
            try:
                tracking = _get_email_tracking(session, email_tracking_id)
                if not tracking:
                    print(f"[quote-pipeline] #{email_tracking_id}: email not found in DB", flush=True)
                    logger.warning(f"[quote-pipeline] #{email_tracking_id}: email not found in DB")
                    return
                # Must be received, linked to RFQ, and have a supplier
                if tracking.direction != "received":
                    print(f"[quote-pipeline] #{email_tracking_id}: skipped (direction={tracking.direction})", flush=True)
                    logger.info(f"[quote-pipeline] #{email_tracking_id}: skipped (direction={tracking.direction})")
                    return
                rfq_id = tracking.rfq_token or tracking.rfq_id
                if not rfq_id:
                    print(f"[quote-pipeline] #{email_tracking_id}: no RFQ link", flush=True)
                    logger.warning(f"[quote-pipeline] #{email_tracking_id}: no RFQ link")
                    return
                if not tracking.supplier_id:
                    print(f"[quote-pipeline] #{email_tracking_id}: no supplier link (rfq_token={tracking.rfq_token})", flush=True)
                    logger.warning(f"[quote-pipeline] #{email_tracking_id}: no supplier link (rfq_token={tracking.rfq_token})")
                    return
            finally:
                session.close()

            # Stage 1: Classify
            classification = _classify_supplier_email_sync(email_tracking_id)

            if classification["classification"] not in ("quote_response", "declined"):
                # Store classification result for non-actionable emails
                _save_pipeline_result(email_tracking_id, {
                    "classification": classification["classification"],
                    "reason": classification["reason"],
                    "supplier_name": classification.get("supplier_name"),
                    "rfq_id": classification.get("rfq_id"),
                    "quote_attachments": classification.get("quote_attachments", []),
                    "processed_at": _now_iso(),
                })
                print(f"[quote-pipeline] #{email_tracking_id}: classified as {classification['classification']}: {classification['reason']}", flush=True)
                logger.info(
                    f"[quote-pipeline] #{email_tracking_id}: classified as "
                    f"{classification['classification']}: {classification['reason']}"
                )
                return

            # Handle declined — mark all supplier items as declined
            if classification["classification"] == "declined":
                supplier_name = classification.get("supplier_name") or "Unknown"
                rfq_id_val = classification.get("rfq_id")
                actions = []
                if rfq_id_val and supplier_name != "Unknown":
                    from includes.tools.rfq_crud import _update_supplier_sync, _get_rfq_dict_sync
                    rfq = _get_rfq_dict_sync(rfq_id_val)
                    if rfq:
                        for item in rfq.get("items", []):
                            for s in (item.get("suppliers") or []):
                                if s.get("name", "").lower() == supplier_name.lower():
                                    _update_supplier_sync(rfq_id_val, {
                                        "line": item["line"],
                                        "name": supplier_name,
                                        "quote_status": "declined",
                                    }, user_id)
                                    actions.append(f"✗ Line {item['line']}: marked declined")
                                    break

                _save_pipeline_result(email_tracking_id, {
                    "classification": "declined",
                    "reason": classification["reason"],
                    "supplier_name": supplier_name,
                    "rfq_id": rfq_id_val,
                    "actions": actions,
                    "processed_at": _now_iso(),
                })
                logger.info(
                    f"[quote-pipeline] #{email_tracking_id}: supplier declined — "
                    f"{len(actions)} items marked"
                )
                return

            # Stage 2: Extract
            content = _extract_email_content_sync(email_tracking_id, classification.get("quote_attachments"))
            if content.startswith("Error:"):
                _save_pipeline_result(email_tracking_id, {
                    "classification": "quote_response",
                    "reason": classification["reason"],
                    "error": content,
                    "processed_at": _now_iso(),
                })
                logger.warning(f"[quote-pipeline] #{email_tracking_id}: extraction failed — {content}")
                return

            # Stage 3: Interpret + Apply
            supplier_name = classification.get("supplier_name") or "Unknown"
            quote_data = _interpret_quote_sync(rfq_id, supplier_name, content)
            if "error" in quote_data:
                _save_pipeline_result(email_tracking_id, {
                    "classification": "quote_response",
                    "reason": classification["reason"],
                    "supplier_name": supplier_name,
                    "error": quote_data["error"],
                    "processed_at": _now_iso(),
                })
                logger.warning(f"[quote-pipeline] #{email_tracking_id}: interpretation failed — {quote_data['error']}")
                return

            actions = _apply_quote_data(rfq_id, supplier_name, quote_data, user_id)

            # Store full pipeline result
            _save_pipeline_result(email_tracking_id, {
                "classification": "quote_response",
                "reason": classification["reason"],
                "supplier_name": supplier_name,
                "rfq_id": rfq_id,
                "quotes": quote_data.get("quotes", []),
                "declined_items": quote_data.get("declined_items", []),
                "shipping": quote_data.get("shipping"),
                "terms": quote_data.get("terms"),
                "notes": quote_data.get("notes"),
                "warnings": quote_data.get("warnings", []),
                "actions": actions,
                "processed_at": _now_iso(),
            })

            logger.info(
                f"[quote-pipeline] #{email_tracking_id}: applied {len(actions)} updates "
                f"to {rfq_id} from {supplier_name}"
            )
        except Exception as e:
            print(f"[quote-pipeline] #{email_tracking_id}: CRASHED — {e}", flush=True)
            logger.error(f"[quote-pipeline] #{email_tracking_id}: failed — {e}", exc_info=True)

    threading.Thread(target=_run, daemon=True, name=f"quote-pipeline-{email_tracking_id}").start()


# ---------------------------------------------------------------------------
# Tool factory
# ---------------------------------------------------------------------------

def create_supplier_quote_tools(user_id: str) -> list:
    """Create supplier quote pipeline tools bound to a user.

    Returns list of tools: [classify_supplier_email, extract_email_content, interpret_quote_response]
    """

    @tool
    async def classify_supplier_email(email_tracking_id: int) -> str:
        """Classify whether an email is a supplier quote response.

        Checks the email body and attachments for pricing indicators,
        verifies it's linked to an RFQ, and identifies the sender supplier.

        Returns classification: 'quote_response', 'not_quote', or 'needs_review',
        along with the reason and matched supplier name.

        Args:
            email_tracking_id: The ID from the email_tracking table.
        """
        result = await asyncio.to_thread(_classify_supplier_email_sync, email_tracking_id)

        lines = [
            f"**Classification:** {result['classification']}",
            f"**Reason:** {result['reason']}",
        ]
        if result.get("rfq_id"):
            lines.append(f"**RFQ:** {result['rfq_id']}")
        if result.get("supplier_name"):
            lines.append(f"**Supplier:** {result['supplier_name']}")

        return "\n".join(lines)

    @tool
    async def extract_email_content(email_tracking_id: int) -> str:
        """Extract full content from an email including attachments.

        Fetches the email body and processes PDF/image attachments using
        Gemini vision to extract pricing tables and text content.
        Returns a Markdown content bundle ready for interpretation.

        Use this after classify_supplier_email confirms it's a quote response.

        Args:
            email_tracking_id: The ID from the email_tracking table.
        """
        return await asyncio.to_thread(_extract_email_content_sync, email_tracking_id)

    @tool
    async def interpret_quote_response(
        rfq_id: str,
        supplier_name: str,
        content_bundle: str,
    ) -> str:
        """Interpret extracted email content and apply quote data to an RFQ.

        Uses the LLM to match pricing from the email content to RFQ items,
        then automatically updates the RFQ with extracted prices, shipping,
        and terms. Returns a summary of what was applied.

        Args:
            rfq_id: The RFQ identifier (e.g. 'RFQ-2026-0039').
            supplier_name: The supplier's name.
            content_bundle: The content bundle from extract_email_content.
        """
        # Step 1: Interpret — extract structured data from content
        quote_data = await asyncio.to_thread(
            _interpret_quote_sync, rfq_id, supplier_name, content_bundle
        )

        if "error" in quote_data:
            return f"❌ Interpretation failed: {quote_data['error']}"

        # Step 2: Apply — write to RFQ
        actions = await asyncio.to_thread(
            _apply_quote_data, rfq_id, supplier_name, quote_data, user_id
        )

        # Step 3: Format summary
        warnings = quote_data.get("warnings", [])
        summary_parts = [
            f"## Quote Applied: {supplier_name} → {rfq_id}",
            "",
        ]

        if actions:
            summary_parts.append("### Updates")
            for a in actions:
                summary_parts.append(f"- {a}")
        else:
            summary_parts.append("*No pricing data could be applied.*")

        if warnings:
            summary_parts.append("")
            summary_parts.append("### Warnings")
            for w in warnings:
                summary_parts.append(f"- ⚠️ {w}")

        return "\n".join(summary_parts)

    return [classify_supplier_email, extract_email_content, interpret_quote_response]
