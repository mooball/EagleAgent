"""Tests for includes/email_pipeline.py — shared email pipeline infrastructure."""

import io
import os
from unittest.mock import MagicMock, patch

import pytest


# ---------------------------------------------------------------------------
# get_pipeline_model — env var resolution order
# ---------------------------------------------------------------------------

class TestGetPipelineModel:
    def test_step_specific_env_var(self, monkeypatch):
        monkeypatch.setenv("QUOTE_CLASSIFY_MODEL", "gemini-custom")
        from includes.email_pipeline import get_pipeline_model
        assert get_pipeline_model("QUOTE", "classify") == "gemini-custom"

    def test_pipeline_level_env_var(self, monkeypatch):
        monkeypatch.delenv("QUOTE_CLASSIFY_MODEL", raising=False)
        monkeypatch.setenv("QUOTE_PIPELINE_MODEL", "gemini-pipeline")
        from includes.email_pipeline import get_pipeline_model
        assert get_pipeline_model("QUOTE", "classify") == "gemini-pipeline"

    def test_falls_back_to_default_model(self, monkeypatch):
        monkeypatch.delenv("QUOTE_CLASSIFY_MODEL", raising=False)
        monkeypatch.delenv("QUOTE_PIPELINE_MODEL", raising=False)
        from includes.email_pipeline import get_pipeline_model
        from config.settings import Config
        assert get_pipeline_model("QUOTE", "classify") == Config.DEFAULT_MODEL

    def test_step_overrides_pipeline(self, monkeypatch):
        monkeypatch.setenv("QUOTE_CLASSIFY_MODEL", "step-model")
        monkeypatch.setenv("QUOTE_PIPELINE_MODEL", "pipeline-model")
        from includes.email_pipeline import get_pipeline_model
        assert get_pipeline_model("QUOTE", "classify") == "step-model"

    def test_different_pipelines_independent(self, monkeypatch):
        monkeypatch.setenv("QUOTE_PIPELINE_MODEL", "quote-model")
        monkeypatch.delenv("CUSTOMER_REQUEST_PIPELINE_MODEL", raising=False)
        monkeypatch.delenv("CUSTOMER_REQUEST_EXTRACT_MODEL", raising=False)
        monkeypatch.delenv("QUOTE_EXTRACT_MODEL", raising=False)
        from includes.email_pipeline import get_pipeline_model
        from config.settings import Config
        assert get_pipeline_model("QUOTE", "extract") == "quote-model"
        assert get_pipeline_model("CUSTOMER_REQUEST", "extract") == Config.DEFAULT_MODEL


# ---------------------------------------------------------------------------
# extract_spreadsheet_content — CSV and Excel parsing
# ---------------------------------------------------------------------------

class TestExtractSpreadsheetContent:
    def test_csv_basic(self):
        from includes.email_pipeline import extract_spreadsheet_content
        csv_data = b"Name,Price,Qty\nWidget,10.50,100\nGadget,25.00,50"
        result = extract_spreadsheet_content(csv_data, "prices.csv", "text/csv")
        assert "```csv" in result
        assert "Widget" in result
        assert "10.50" in result

    def test_csv_truncation(self):
        from includes.email_pipeline import extract_spreadsheet_content
        long_csv = b"x" * 6000
        result = extract_spreadsheet_content(long_csv, "big.csv", "text/csv")
        assert len(result) <= 5100  # 5000 + markdown fencing

    def test_xlsx_basic(self):
        from includes.email_pipeline import extract_spreadsheet_content
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pricing"
        ws.append(["Part", "Price", "Lead Time"])
        ws.append(["ABC-123", 42.50, "4 weeks"])
        ws.append(["DEF-456", 18.00, "2 weeks"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        result = extract_spreadsheet_content(xlsx_bytes, "quote.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert "## Sheet: Pricing" in result
        assert "ABC-123" in result
        assert "42.5" in result
        assert "4 weeks" in result

    def test_xlsx_empty_sheet(self):
        from includes.email_pipeline import extract_spreadsheet_content
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        buf = io.BytesIO()
        wb.save(buf)
        result = extract_spreadsheet_content(buf.getvalue(), "empty.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert "empty" in result.lower()

    def test_xlsx_none_cells(self):
        from includes.email_pipeline import extract_spreadsheet_content
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["A", None, "C"])
        ws.append([None, "B2", None])
        buf = io.BytesIO()
        wb.save(buf)
        result = extract_spreadsheet_content(buf.getvalue(), "sparse.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert "B2" in result

    def test_xlsx_multi_sheet(self):
        from includes.email_pipeline import extract_spreadsheet_content
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Col1"])
        ws1.append(["Val1"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Col2"])
        ws2.append(["Val2"])
        buf = io.BytesIO()
        wb.save(buf)
        result = extract_spreadsheet_content(buf.getvalue(), "multi.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert "Sheet1" in result
        assert "Sheet2" in result
        assert "Val1" in result
        assert "Val2" in result

    def test_xlsx_truncation(self):
        from includes.email_pipeline import extract_spreadsheet_content
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Data"])
        for i in range(300):
            ws.append([f"Row {i} with lots of text to make it larger" * 5])
        buf = io.BytesIO()
        wb.save(buf)
        result = extract_spreadsheet_content(buf.getvalue(), "big.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert len(result) <= 8000

    def test_corrupt_file(self):
        from includes.email_pipeline import extract_spreadsheet_content
        result = extract_spreadsheet_content(b"not a real xlsx", "bad.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert "failed" in result.lower()


# ---------------------------------------------------------------------------
# llm_call_with_retry — retry and fallback logic
# ---------------------------------------------------------------------------

class TestLlmCallWithRetry:
    def test_success_first_attempt(self):
        from includes.email_pipeline import llm_call_with_retry

        mock_response = MagicMock()
        mock_response.text = "result"
        mock_client = MagicMock()
        mock_client.models.generate_content.return_value = mock_response

        with patch("includes.email_pipeline.get_pipeline_model", return_value="test-model"), \
             patch("google.genai.Client", return_value=mock_client):
            result = llm_call_with_retry("QUOTE", "classify", ["test"])
            assert result.text == "result"
            assert mock_client.models.generate_content.call_count == 1

    def test_retries_on_transient_error(self):
        from includes.email_pipeline import llm_call_with_retry

        mock_response = MagicMock()
        mock_response.text = "ok"
        mock_client = MagicMock()
        mock_client.models.generate_content.side_effect = [
            Exception("504 Deadline Exceeded"),
            mock_response,
        ]

        with patch("includes.email_pipeline.get_pipeline_model", return_value="test-model"), \
             patch("google.genai.Client", return_value=mock_client), \
             patch("time.sleep"):
            result = llm_call_with_retry("QUOTE", "classify", ["test"])
            assert result.text == "ok"
            assert mock_client.models.generate_content.call_count == 2

    def test_falls_back_to_fallback_model(self):
        from includes.email_pipeline import llm_call_with_retry, FALLBACK_MODEL

        mock_response = MagicMock()
        mock_response.text = "fallback ok"
        mock_client = MagicMock()
        mock_client.models.generate_content.side_effect = [
            Exception("503 Service Unavailable"),
            Exception("504 Gateway Timeout"),
            mock_response,
        ]

        with patch("includes.email_pipeline.get_pipeline_model", return_value="primary"), \
             patch("google.genai.Client", return_value=mock_client), \
             patch("time.sleep"):
            result = llm_call_with_retry("QUOTE", "classify", ["test"])
            assert result.text == "fallback ok"
            # Third call should use FALLBACK_MODEL
            calls = mock_client.models.generate_content.call_args_list
            assert calls[2][1]["model"] == FALLBACK_MODEL

    def test_permanent_error_not_retried(self):
        from includes.email_pipeline import llm_call_with_retry

        mock_client = MagicMock()
        mock_client.models.generate_content.side_effect = ValueError("Invalid input")

        with patch("includes.email_pipeline.get_pipeline_model", return_value="test-model"), \
             patch("google.genai.Client", return_value=mock_client):
            with pytest.raises(ValueError, match="Invalid input"):
                llm_call_with_retry("QUOTE", "classify", ["test"])
            assert mock_client.models.generate_content.call_count == 1

    def test_all_retries_exhausted(self):
        from includes.email_pipeline import llm_call_with_retry

        mock_client = MagicMock()
        mock_client.models.generate_content.side_effect = Exception("503 UNAVAILABLE")

        with patch("includes.email_pipeline.get_pipeline_model", return_value="test-model"), \
             patch("google.genai.Client", return_value=mock_client), \
             patch("time.sleep"):
            with pytest.raises(Exception, match="503"):
                llm_call_with_retry("QUOTE", "classify", ["test"])
            assert mock_client.models.generate_content.call_count == 3


# ---------------------------------------------------------------------------
# Image signature caching
# ---------------------------------------------------------------------------

class TestImageSignature:
    def test_check_known_signature(self):
        from includes.email_pipeline import check_image_signature

        mock_record = MagicMock()
        mock_record.classification = "signature"
        mock_session = MagicMock()
        mock_session.query.return_value.filter.return_value.first.return_value = mock_record

        with patch("includes.email_pipeline._get_session", return_value=mock_session):
            result = check_image_signature(b"test image bytes")
            assert result == "signature"

    def test_check_unknown_signature(self):
        from includes.email_pipeline import check_image_signature

        mock_session = MagicMock()
        mock_session.query.return_value.filter.return_value.first.return_value = None

        with patch("includes.email_pipeline._get_session", return_value=mock_session):
            result = check_image_signature(b"new image bytes")
            assert result is None

    def test_store_signature(self):
        from includes.email_pipeline import store_image_signature
        import hashlib

        mock_session = MagicMock()
        # Return None for the "existing" query so the function proceeds to add
        mock_session.query.return_value.filter.return_value.first.return_value = None
        image_bytes = b"test image"
        expected_sha = hashlib.sha256(image_bytes).hexdigest()

        with patch("includes.email_pipeline._get_session", return_value=mock_session):
            store_image_signature(image_bytes, "signature", "logo.png", 42)
            mock_session.add.assert_called_once()
            added = mock_session.add.call_args[0][0]
            assert added.sha256 == expected_sha
            assert added.classification == "signature"
            assert added.sample_filename == "logo.png"
            mock_session.commit.assert_called_once()

    def test_store_signature_skips_duplicate(self):
        from includes.email_pipeline import store_image_signature

        mock_existing = MagicMock()
        mock_session = MagicMock()
        mock_session.query.return_value.filter.return_value.first.return_value = mock_existing

        with patch("includes.email_pipeline._get_session", return_value=mock_session):
            store_image_signature(b"test image", "signature")
            mock_session.add.assert_not_called()
